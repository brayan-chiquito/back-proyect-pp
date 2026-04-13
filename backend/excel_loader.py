"""Descarga y parseo del Excel de catálogo (openpyxl + caché en memoria)."""

from __future__ import annotations

import asyncio
import logging
import re
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Any

import httpx
from openpyxl import load_workbook

from config import get_settings

logger = logging.getLogger(__name__)

_TENANT_SHEETS: dict[str, str] = {
    "pasteleria": "Pasteleria",
    "pijamas": "Pijamas",
    "comida": "Comida",
}

_products_by_tenant: dict[str, list[dict[str, Any]]] = {}
_loaded_at: datetime | None = None
_lock = asyncio.Lock()


def _norm_header(cell: object) -> str:
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell).strip().lower())


def _canon_field(name: str) -> str | None:
    m = {
        "nombre": "nombre",
        "descripcion": "descripcion",
        "descripción": "descripcion",
        "precio": "precio",
        "disponible": "disponible",
    }
    return m.get(name)


def _parse_boolish(val: object) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    s = str(val).strip().lower()
    return s in ("sí", "si", "s", "yes", "true", "1", "x", "ok", "y")


def _parse_sheet(ws: Any) -> list[dict[str, Any]]:
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return []
    idx_map: dict[int, str] = {}
    for i, cell in enumerate(header_row):
        canon = _canon_field(_norm_header(cell))
        if canon:
            idx_map[i] = canon
    need = {"nombre", "descripcion", "precio", "disponible"}
    if not need.issubset(set(idx_map.values())):
        logger.warning(
            "excel_loader: faltan columnas esperadas en %s",
            ws.title,
        )
    out: list[dict[str, Any]] = []
    for row in rows:
        if row is None or all(v is None for v in row):
            continue
        item: dict[str, Any] = {}
        for col_idx, field in idx_map.items():
            if col_idx < len(row):
                item[field] = row[col_idx]
        raw_nombre = item.get("nombre")
        if raw_nombre is None or str(raw_nombre).strip() == "":
            continue
        out.append(item)
    return out


def _parse_workbook_sync(content: bytes) -> dict[str, list[dict[str, Any]]]:
    bio = BytesIO(content)
    wb = load_workbook(bio, read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    try:
        names = set(wb.sheetnames)
        for slug, title in _TENANT_SHEETS.items():
            if title not in names:
                logger.warning("excel_loader: no existe hoja %r", title)
                result[slug] = []
                continue
            result[slug] = _parse_sheet(wb[title])
    finally:
        wb.close()
    return result


async def _download_and_parse(url: str) -> dict[str, list[dict[str, Any]]]:
    timeout = httpx.Timeout(60.0, connect=15.0)
    async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
        resp = await client.get(url)
        resp.raise_for_status()
    content = resp.content
    return await asyncio.to_thread(_parse_workbook_sync, content)


def format_products_for_context(products: list[dict[str, Any]]) -> str:
    """Texto plano para system message (solo ítems marcados disponibles)."""
    lines = ["Productos disponibles:"]
    any_row = False
    for p in products:
        if not _parse_boolish(p.get("disponible")):
            continue
        any_row = True
        nombre = str(p.get("nombre", "")).strip() or "—"
        precio = p.get("precio")
        if isinstance(precio, (int, float)):
            precio_s = f"${precio:,.0f}".replace(",", ".")
        else:
            precio_s = str(precio).strip() or "—"
        desc = str(p.get("descripcion", "")).strip().replace("\n", " ") or "—"
        lines.append(
            f"- Nombre: {nombre} | Precio: {precio_s} | Desc: {desc}",
        )
    if not any_row:
        lines.append("(no hay productos disponibles en catálogo)")
    return "\n".join(lines)


async def load_products(tenant: str) -> list[dict[str, Any]]:
    """
    Lee EXCEL_URL (un .xlsx con hojas Pasteleria / Pijamas / Comida).
    Caché en memoria EXCEL_CACHE_MINUTES; si falla la URL, último caché.
    """
    global _loaded_at
    key = tenant.strip().lower()
    if key not in _TENANT_SHEETS:
        logger.warning("excel_loader: tenant desconocido %r", tenant)
        return []

    settings = get_settings()
    url = settings.excel_url.strip()
    ttl_min = max(1, settings.excel_cache_minutes)
    now = datetime.now(timezone.utc)

    async with _lock:
        fresh = False
        if _loaded_at is not None and _products_by_tenant:
            age = now - _loaded_at
            fresh = age < timedelta(minutes=ttl_min)
        if fresh:
            return list(_products_by_tenant.get(key, []))

        try:
            if not url:
                raise ValueError("EXCEL_URL vacía")
            parsed = await _download_and_parse(url)
            _products_by_tenant.clear()
            _products_by_tenant.update(parsed)
            _loaded_at = now
            logger.info("excel_loader: catálogo cargado desde URL")
        except Exception as exc:
            logger.error(
                "excel_loader: error al obtener o parsear Excel: %s",
                exc,
                exc_info=True,
            )
            if _products_by_tenant:
                logger.warning(
                    "excel_loader: se mantiene el último caché en memoria",
                )
            else:
                logger.warning("excel_loader: no hay caché previo")

        return list(_products_by_tenant.get(key, []))
